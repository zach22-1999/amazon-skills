#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
zach-product-research 交付生成/校验脚本 v2

输入一个统一 JSON 数据包，生成：
1. Markdown 正式报告（v2: 从 chapters 模板渲染）
2. HTML 精简报告（v2: 从 chapters 模板渲染）
3. Excel 多 Sheet 数据文件
4. Dashboard 可视化看板（v2 新增）
5. 原始/中间 JSON 工件

v2 核心变化：
- chapters 结构化数据 + LLM 洞察段落 = 唯一事实来源
- MD/HTML 由渲染模板生成表格 + 插入 LLM 洞察
- Dashboard 从 chapters 直接构建 view-model（替代 data_loader.py 1945 行猜测性解析）

向后兼容：检测到 schema_version != "2.0" 或有 report_markdown 字段时，走 v1 旧路径。

用法：
  python render_deliverables.py generate --input payload.json
  python render_deliverables.py validate --input payload.json
  python render_deliverables.py all --input payload.json

  # 从任意目录运行（使用绝对路径）：
  python render_deliverables.py all --input /absolute/path/to/payload.json --root /path/to/repo/root

  # 或使用相对路径（相对于脚本所在目录）：
  cd /path/to/skills/zach-product-research/scripts
  python render_deliverables.py all --input ../payload.json
"""

from __future__ import annotations

import argparse
import html as html_lib
import json
import re
import sys
from pathlib import Path
from typing import Any

# Windows 编码兼容性：确保 UTF-8 输出正常
if sys.platform == "win32":
    import io
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REQUIRED_METADATA = [
    "brand", "category", "version", "date", "site",
    "keyword", "topic", "type", "data_sources",
]

REQUIRED_SHEETS = [
    "数据来源说明", "市场概况", "类目销量Top100_明细",
    "关键词对比_分段", "新品分析",
    "竞品选择逻辑", "竞品差评摘要", "品牌_竞品格局",
    "属性标注_Top100", "进入壁垒评估", "Go-NoGo评分卡",
]

CHAPTER_KEYS = [
    "ch01_executive_summary",
    "ch02_market_overview",
    "ch03_dimension_distribution",
    "ch04_cross_analysis",
    "ch05_competitor_brands",
    "ch06_voc_pain_points",
    "ch07_opportunity_gaps",
    "ch08_strategic_recommendations",
    "ch09_barriers_go_nogo",
    "ch10_appendix",
]

CHAPTER_TITLES = {
    "ch01_executive_summary": "一、Executive Summary（核心摘要）",
    "ch02_market_overview": "二、市场总览",
    "ch03_dimension_distribution": "三、N 维产品分布",
    "ch04_cross_analysis": "四、交叉分析",
    "ch05_competitor_brands": "五、竞品与品牌格局",
    "ch06_voc_pain_points": "六、VOC 消费者痛点",
    "ch07_opportunity_gaps": "七、机会空白汇总",
    "ch08_strategic_recommendations": "八、品牌战略建议",
    "ch09_barriers_go_nogo": "九、进入壁垒与 Go/No-Go 评分",
    "ch10_appendix": "十、数据来源与附录",
}


def _is_v2(payload: dict[str, Any]) -> bool:
    return payload.get("schema_version") == "2.0" and "chapters" in payload


# ---------------------------------------------------------------------------
# Payload reading
# ---------------------------------------------------------------------------

def read_payload(path: Path) -> dict[str, Any]:
    """读取 JSON payload，自动处理 Windows/macOS 的编码差异"""
    try:
        # UTF-8 with BOM 和不带 BOM 都支持
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except UnicodeDecodeError:
        # 备用编码
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except UnicodeDecodeError:
            raise ValueError(f"无法以 UTF-8 编码读取文件：{path}。请确保文件编码为 UTF-8。")

    if not isinstance(payload, dict):
        raise ValueError("payload 顶层必须是 JSON object")
    metadata = payload.get("metadata")
    if not isinstance(metadata, dict):
        raise ValueError("payload.metadata 缺失或格式错误")
    missing = [key for key in REQUIRED_METADATA if key not in metadata]
    if missing:
        raise ValueError(f"metadata 缺少字段: {', '.join(missing)}")
    if not isinstance(metadata["data_sources"], list) or not metadata["data_sources"]:
        raise ValueError("metadata.data_sources 必须是非空列表")
    if _is_v2(payload):
        if "excel_sheets" not in payload or not isinstance(payload["excel_sheets"], dict):
            raise ValueError("v2 payload 必须包含 excel_sheets (非空 object)")
    else:
        if "report_markdown" not in payload or "report_html" not in payload or "excel_sheets" not in payload:
            raise ValueError("v1 payload 必须包含 report_markdown / report_html / excel_sheets")
        if not isinstance(payload["excel_sheets"], dict) or not payload["excel_sheets"]:
            raise ValueError("excel_sheets 必须是非空 object")
    return payload


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

def build_output_dir(metadata: dict[str, Any], root: Path) -> Path:
    return root / "工作成果" / "brands" / metadata["brand"] / "市场调研" / metadata["category"] / metadata["version"]


def build_base_name(metadata: dict[str, Any]) -> str:
    return f'{metadata["date"]}_{metadata["site"]}_{metadata["keyword"]}'


def build_paths(metadata: dict[str, Any], root: Path) -> dict[str, Path]:
    output_dir = build_output_dir(metadata, root)
    base_name = build_base_name(metadata)
    version = metadata["version"]
    return {
        "dir": output_dir,
        "md": output_dir / f"{base_name}_市场调研报告_{version}.md",
        "html": output_dir / f"{base_name}_精简报告_{version}.html",
        "xlsx": output_dir / f"{base_name}_市场调研_数据_{version}.xlsx",
        "dashboard": output_dir / f"{base_name}_可视化看板_{version}.html",
    }


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def render_frontmatter(metadata: dict[str, Any]) -> str:
    sources = "\n".join(f"  - {item}" for item in metadata["data_sources"])
    created = metadata.get("created", "")
    return (
        "---\n"
        f"created: {created}\n"
        f'topic: {metadata["topic"]}\n'
        f'type: {metadata["type"]}\n'
        "data_sources:\n"
        f"{sources}\n"
        "---\n\n"
    )


def _md_table(headers: list[str], rows: list[list[str]]) -> str:
    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    for row in rows:
        lines.append("| " + " | ".join(str(c) for c in row) + " |")
    return "\n".join(lines)


def _safe(val: Any) -> str:
    if val is None:
        return ""
    return str(val)


# ---------------------------------------------------------------------------
# v2 Markdown renderer
# ---------------------------------------------------------------------------

def render_markdown(chapters: dict[str, Any], metadata: dict[str, Any]) -> str:
    parts: list[str] = []
    parts.append(render_frontmatter(metadata))
    topic = metadata.get("topic", "选品分析报告")
    parts.append(f"# {topic}\n")

    for ch_key in CHAPTER_KEYS:
        ch = chapters.get(ch_key, {})
        title = CHAPTER_TITLES[ch_key]
        parts.append(f"\n## {title}\n")
        renderer = _MD_RENDERERS.get(ch_key)
        if renderer:
            parts.append(renderer(ch))

    return "\n".join(parts)


def _render_ch01_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    verdict = ch.get("go_nogo_verdict", "")
    summary = ch.get("one_line_summary", "")
    if verdict:
        lines.append(f"**决策建议：{verdict}**\n")
    if summary:
        lines.append(f"> {summary}\n")
    conclusions = ch.get("conclusions", [])
    if conclusions:
        lines.append("### 核心结论\n")
        for i, c in enumerate(conclusions, 1):
            dp = _safe(c.get("data_point"))
            meaning = _safe(c.get("meaning"))
            action = _safe(c.get("action"))
            lines.append(f"{i}. **{dp}** → {meaning} → 💡 {action}")
    return "\n".join(lines)


def _render_ch02_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    kpis = ch.get("kpis", [])
    if kpis:
        headers = ["指标", "数值", "来源"]
        rows = [[_safe(k.get("name")), _safe(k.get("value")), _safe(k.get("source"))] for k in kpis]
        lines.append(_md_table(headers, rows))
    kw = ch.get("keyword_comparison", [])
    if kw:
        lines.append("\n### 关键词对比\n")
        headers = ["关键词", "搜索量排名", "搜索量变化", "分段"]
        rows = [
            [_safe(k.get("keyword")), _safe(k.get("search_volume_rank")),
             _safe(k.get("search_volume_change")), _safe(k.get("segment"))]
            for k in kw
        ]
        lines.append(_md_table(headers, rows))
    insight = ch.get("insight", "")
    if insight:
        lines.append(f"\n**关键洞察：**\n\n{insight}")
    return "\n".join(lines)


def _render_ch03_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    for dim in ch.get("dimensions", []):
        lines.append(f"\n### {_safe(dim.get('name'))}\n")
        dist = dim.get("distribution", [])
        if dist:
            headers = ["维度值", "产品数", "占比", "月销量", "销量占比", "月销额", "均价"]
            rows = [
                [_safe(d.get("value")), _safe(d.get("product_count")), _safe(d.get("product_share")),
                 _safe(d.get("monthly_sales")), _safe(d.get("sales_share")),
                 _safe(d.get("monthly_revenue")), _safe(d.get("avg_price"))]
                for d in dist
            ]
            lines.append(_md_table(headers, rows))
        insight = dim.get("insight", "")
        if insight:
            lines.append(f"\n**关键洞察：**\n\n{insight}")
    return "\n".join(lines)


def _render_ch04_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    for m in ch.get("matrices", []):
        d1, d2 = _safe(m.get("dim1")), _safe(m.get("dim2"))
        lines.append(f"\n### {d1} × {d2}\n")
        data = m.get("data", [])
        if data:
            headers = [d1, d2, "产品数", "月销量", "月销额"]
            rows = [
                [_safe(d.get("dim1_value")), _safe(d.get("dim2_value")),
                 _safe(d.get("count")), _safe(d.get("sales")), _safe(d.get("revenue"))]
                for d in data
            ]
            lines.append(_md_table(headers, rows))
        gaps = m.get("gaps", [])
        if gaps:
            lines.append("\n**供需缺口：**\n")
            for g in gaps:
                lines.append(f"- {_safe(g.get('dim1_value'))} × {_safe(g.get('dim2_value'))}（产品数 {_safe(g.get('count'))}）：{_safe(g.get('reason'))}")
        findings = m.get("findings", "")
        if findings:
            lines.append(f"\n**关键发现：**\n\n{findings}")
    return "\n".join(lines)


def _render_ch05_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    logic = ch.get("competitor_selection_logic", [])
    if logic:
        lines.append("### 竞品选择逻辑\n")
        headers = ["ASIN", "品牌", "选择理由", "竞品类型", "覆盖维度"]
        rows = [
            [_safe(c.get("asin")), _safe(c.get("brand")), _safe(c.get("reason")),
             _safe(c.get("type")), _safe(c.get("covered_dimensions"))]
            for c in logic
        ]
        lines.append(_md_table(headers, rows))
    landscape = ch.get("brand_landscape", [])
    if landscape:
        lines.append("\n### 品牌格局\n")
        headers = ["品牌", "市场份额", "产品数", "均价", "策略", "角色", "对我方含义"]
        rows = [
            [_safe(b.get("brand")), _safe(b.get("market_share")), _safe(b.get("product_count")),
             _safe(b.get("avg_price")), _safe(b.get("strategy")),
             _safe(b.get("role")), _safe(b.get("implication_for_us"))]
            for b in landscape
        ]
        lines.append(_md_table(headers, rows))
    insight = ch.get("insight", "")
    if insight:
        lines.append(f"\n**关键洞察：**\n\n{insight}")
    return "\n".join(lines)


def _render_ch06_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    for dim in ch.get("by_dimension", []):
        lines.append(f"\n### {_safe(dim.get('dimension'))}\n")
        pps = dim.get("pain_points", [])
        if pps:
            headers = ["痛点描述", "频次", "占比", "涉及竞品"]
            rows = [
                [_safe(p.get("description")), _safe(p.get("frequency")),
                 _safe(p.get("share")), _safe(p.get("related_competitors"))]
                for p in pps
            ]
            lines.append(_md_table(headers, rows))
        support = dim.get("support_data", "")
        if support:
            lines.append(f"\n**数据支撑：** {support}")

        lines.append("\n| 痛点描述 | 数据支撑 | 品牌机会 | 产品方案 |")
        lines.append("| --- | --- | --- | --- |")
        opp = _safe(dim.get("opportunity"))
        sol = _safe(dim.get("solution"))
        for p in pps[:1]:
            lines.append(f"| {_safe(p.get('description'))} | {support} | {opp} | {sol} |")
    return "\n".join(lines)


def _render_ch07_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    method = ch.get("scoring_method", "")
    if method:
        lines.append(f"**评分方法：** {method}\n")
    opps = ch.get("opportunities", [])
    if opps:
        headers = ["优先级", "机会描述", "市场规模", "可行性", "品牌匹配", "加权得分", "行动建议"]
        rows = [
            [_safe(o.get("rank")), _safe(o.get("description")),
             _safe(o.get("market_size_score")), _safe(o.get("feasibility_score")),
             _safe(o.get("brand_fit_score")), _safe(o.get("weighted_score")),
             _safe(o.get("action"))]
            for o in opps
        ]
        lines.append(_md_table(headers, rows))
    insight = ch.get("insight", "")
    if insight:
        lines.append(f"\n**关键洞察：**\n\n{insight}")
    return "\n".join(lines)


def _render_ch08_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    for t in ch.get("tiers", []):
        tier_name = _safe(t.get("tier"))
        product_name = _safe(t.get("name", ""))
        lines.append(f"\n### {tier_name}：{product_name}\n")
        specs = t.get("specs", {})
        if specs:
            headers = ["维度", "规格", "决策依据"]
            rows = [[k, v, ""] for k, v in specs.items()]
            lines.append(_md_table(headers, rows))
        lines.append(f"\n- **目标定价**：{_safe(t.get('target_price'))}")
        lines.append(f"- **差异化主张**：{_safe(t.get('differentiation'))}")
        lines.append(f"- **对标竞品**：{_safe(t.get('benchmark_asin'))}")
        lines.append(f"- **竞品对比优势**：{_safe(t.get('benchmark_advantage'))}")
        lines.append(f"- **预估月销**：{_safe(t.get('estimated_monthly_sales'))}")
        rationale = t.get("decision_rationale", "")
        if rationale:
            lines.append(f"- **决策依据**：{rationale}")
    strategy = ch.get("brand_strategy", "")
    if strategy:
        lines.append(f"\n### 品牌策略\n\n{strategy}")
    timeline = ch.get("timeline", "")
    if timeline:
        lines.append(f"\n### 时间线建议\n\n{timeline}")
    return "\n".join(lines)


def _render_ch09_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    barriers = ch.get("barriers", [])
    if barriers:
        lines.append("### 进入壁垒评估\n")
        headers = ["壁垒类型", "等级", "详情", "应对策略"]
        rows = [
            [_safe(b.get("type")), _safe(b.get("level")),
             _safe(b.get("detail")), _safe(b.get("mitigation"))]
            for b in barriers
        ]
        lines.append(_md_table(headers, rows))
    scorecard = ch.get("go_nogo_scorecard", [])
    if scorecard:
        lines.append("\n### Go/No-Go 评分卡\n")
        headers = ["维度", "权重(%)", "得分", "加权得分", "评分依据"]
        rows = [
            [_safe(s.get("dimension")), _safe(s.get("weight")),
             _safe(s.get("score")), _safe(s.get("weighted_score")),
             _safe(s.get("rationale"))]
            for s in scorecard
        ]
        lines.append(_md_table(headers, rows))
    total = ch.get("total_score")
    verdict = ch.get("verdict", "")
    if total is not None:
        lines.append(f"\n**总分：{total} / 10**")
    if verdict:
        lines.append(f"\n**决策：{verdict}**")
    detail = ch.get("verdict_detail", "")
    if detail:
        lines.append(f"\n{detail}")
    risks = ch.get("risks", [])
    if risks:
        lines.append("\n### 风险提示\n")
        headers = ["风险", "概率", "影响", "应对策略"]
        rows = [
            [_safe(r.get("risk")), _safe(r.get("probability")),
             _safe(r.get("impact")), _safe(r.get("mitigation"))]
            for r in risks
        ]
        lines.append(_md_table(headers, rows))
    return "\n".join(lines)


def _render_ch10_md(ch: dict[str, Any]) -> str:
    lines: list[str] = []
    data_date = ch.get("data_date", "")
    note = ch.get("data_freshness_note", "")
    if data_date:
        lines.append(f"**数据获取日期：** {data_date}")
    if note:
        lines.append(f"\n**时效提醒：** {note}")
    calls = ch.get("tool_calls", [])
    if calls:
        lines.append("\n### Sorftime MCP 工具调用清单\n")
        headers = ["工具", "参数", "日期", "用途"]
        rows = [
            [_safe(c.get("tool")), _safe(c.get("params")),
             _safe(c.get("date")), _safe(c.get("purpose"))]
            for c in calls
        ]
        lines.append(_md_table(headers, rows))
    files = ch.get("files", [])
    if files:
        lines.append("\n### 文件清单\n")
        headers = ["文件名", "类型", "说明"]
        rows = [[_safe(f.get("filename")), _safe(f.get("type")), _safe(f.get("description"))] for f in files]
        lines.append(_md_table(headers, rows))
    return "\n".join(lines)


_MD_RENDERERS = {
    "ch01_executive_summary": _render_ch01_md,
    "ch02_market_overview": _render_ch02_md,
    "ch03_dimension_distribution": _render_ch03_md,
    "ch04_cross_analysis": _render_ch04_md,
    "ch05_competitor_brands": _render_ch05_md,
    "ch06_voc_pain_points": _render_ch06_md,
    "ch07_opportunity_gaps": _render_ch07_md,
    "ch08_strategic_recommendations": _render_ch08_md,
    "ch09_barriers_go_nogo": _render_ch09_md,
    "ch10_appendix": _render_ch10_md,
}


# ---------------------------------------------------------------------------
# v2 HTML 精简报告 renderer
# ---------------------------------------------------------------------------

def _h(text: Any) -> str:
    """HTML-escape helper."""
    return html_lib.escape(str(text)) if text else ""


def render_html_report(chapters: dict[str, Any], metadata: dict[str, Any]) -> str:
    topic = _h(metadata.get("topic", "选品分析报告"))
    created = _h(metadata.get("created", ""))
    sources = ", ".join(metadata.get("data_sources", []))

    ch01 = chapters.get("ch01_executive_summary", {})
    ch02 = chapters.get("ch02_market_overview", {})
    ch03 = chapters.get("ch03_dimension_distribution", {})
    ch05 = chapters.get("ch05_competitor_brands", {})
    ch06 = chapters.get("ch06_voc_pain_points", {})
    ch07 = chapters.get("ch07_opportunity_gaps", {})
    ch08 = chapters.get("ch08_strategic_recommendations", {})
    ch09 = chapters.get("ch09_barriers_go_nogo", {})

    # -- KPI cards --
    kpi_html = ""
    for k in ch02.get("kpis", []):
        kpi_html += (
            f'<div class="kpi-card">'
            f'<div class="label">{_h(k.get("name"))}</div>'
            f'<div class="value">{_h(k.get("value"))}</div>'
            f'</div>\n'
        )

    # -- Executive summary --
    summary_html = ""
    for c in ch01.get("conclusions", []):
        summary_html += (
            f'<div class="conclusion-card">'
            f'<strong>{_h(c.get("data_point"))}</strong> &rarr; '
            f'{_h(c.get("meaning"))} &rarr; '
            f'<em>{_h(c.get("action"))}</em>'
            f'</div>\n'
        )
    verdict = _h(ch01.get("go_nogo_verdict", ch09.get("verdict", "")))
    one_line = _h(ch01.get("one_line_summary", ""))

    # -- Dimension tables --
    dim_html = ""
    for dim in ch03.get("dimensions", []):
        dim_html += f'<h3>{_h(dim.get("name"))}</h3><table><tr><th>维度值</th><th>产品数</th><th>占比</th><th>月销量</th><th>销量占比</th><th>均价</th></tr>'
        for d in dim.get("distribution", []):
            dim_html += (
                f'<tr><td>{_h(d.get("value"))}</td><td>{_h(d.get("product_count"))}</td>'
                f'<td>{_h(d.get("product_share"))}</td><td>{_h(d.get("monthly_sales"))}</td>'
                f'<td>{_h(d.get("sales_share"))}</td><td>{_h(d.get("avg_price"))}</td></tr>'
            )
        dim_html += "</table>"
        ins = dim.get("insight", "")
        if ins:
            dim_html += f'<p class="insight">{_h(ins)}</p>'

    # -- Pain points --
    pain_html = ""
    for dim in ch06.get("by_dimension", []):
        pain_html += f'<div class="pain-card"><h4>{_h(dim.get("dimension"))}</h4>'
        for p in dim.get("pain_points", [])[:3]:
            pain_html += f'<p>{_h(p.get("description"))} (频次 {_h(p.get("frequency"))})</p>'
        opp = dim.get("opportunity", "")
        if opp:
            pain_html += f'<p class="opportunity"><strong>机会：</strong>{_h(opp)}</p>'
        pain_html += "</div>\n"

    # -- Gap table --
    gap_html = "<table><tr><th>#</th><th>机会描述</th><th>加权得分</th><th>行动建议</th></tr>"
    for o in ch07.get("opportunities", []):
        gap_html += (
            f'<tr><td>{_h(o.get("rank"))}</td><td>{_h(o.get("description"))}</td>'
            f'<td>{_h(o.get("weighted_score"))}</td><td>{_h(o.get("action"))}</td></tr>'
        )
    gap_html += "</table>"

    # -- Brand landscape --
    brand_html = ""
    for b in ch05.get("brand_landscape", []):
        brand_html += (
            f'<tr><td>{_h(b.get("brand"))}</td><td>{_h(b.get("market_share"))}</td>'
            f'<td>{_h(b.get("product_count"))}</td><td>{_h(b.get("strategy"))}</td>'
            f'<td>{_h(b.get("implication_for_us"))}</td></tr>'
        )
    if brand_html:
        brand_html = "<table><tr><th>品牌</th><th>份额</th><th>产品数</th><th>策略</th><th>对我方含义</th></tr>" + brand_html + "</table>"

    # -- Strategy cards --
    tier_html = ""
    for t in ch08.get("tiers", []):
        tier_html += (
            f'<div class="tier-card">'
            f'<h4>{_h(t.get("tier"))}：{_h(t.get("name", ""))}</h4>'
            f'<p>定价：{_h(t.get("target_price"))}</p>'
            f'<p>差异化：{_h(t.get("differentiation"))}</p>'
            f'<p>对标：{_h(t.get("benchmark_asin"))}</p>'
            f'</div>\n'
        )

    # -- Barriers --
    barrier_html = "<table><tr><th>壁垒类型</th><th>等级</th><th>详情</th><th>应对</th></tr>"
    for b in ch09.get("barriers", []):
        barrier_html += (
            f'<tr><td>{_h(b.get("type"))}</td><td>{_h(b.get("level"))}</td>'
            f'<td>{_h(b.get("detail"))}</td><td>{_h(b.get("mitigation"))}</td></tr>'
        )
    barrier_html += "</table>"

    # -- Go/No-Go scorecard --
    score_html = "<table><tr><th>维度</th><th>权重</th><th>得分</th><th>加权</th></tr>"
    for s in ch09.get("go_nogo_scorecard", []):
        score_html += (
            f'<tr><td>{_h(s.get("dimension"))}</td><td>{_h(s.get("weight"))}%</td>'
            f'<td>{_h(s.get("score"))}</td><td>{_h(s.get("weighted_score"))}</td></tr>'
        )
    score_html += "</table>"
    total_score = ch09.get("total_score", "")
    verdict_detail = _h(ch09.get("verdict_detail", ""))

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{topic}</title>
  <style>
    :root {{
      --blue: #2563eb; --blue-light: #eff6ff; --green: #16a34a;
      --orange: #ea580c; --red: #dc2626;
      --gray-50: #f9fafb; --gray-100: #f3f4f6; --gray-200: #e5e7eb;
      --gray-500: #6b7280; --gray-700: #374151; --gray-900: #111827;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      color: var(--gray-700); background: var(--gray-50); line-height: 1.6;
    }}
    .container {{ max-width: 960px; margin: 0 auto; padding: 40px 24px 80px; }}
    .header {{ text-align: center; margin-bottom: 48px; }}
    .header h1 {{ margin: 0 0 8px; color: var(--gray-900); font-size: 28px; }}
    .subtitle {{ color: var(--gray-500); font-size: 14px; }}
    .verdict-badge {{
      display: inline-block; padding: 6px 16px; border-radius: 20px;
      font-weight: 700; font-size: 16px; margin: 12px 0;
      background: var(--blue-light); color: var(--blue);
    }}
    section {{ margin-bottom: 40px; }}
    section h2 {{ font-size: 18px; color: var(--gray-900); margin-bottom: 16px;
      padding-bottom: 8px; border-bottom: 2px solid var(--gray-200); }}
    .kpi-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 24px; }}
    .kpi-card, .pain-card, .tier-card, .conclusion-card {{
      background: white; border: 1px solid var(--gray-200); border-radius: 10px;
    }}
    .kpi-card {{ padding: 16px; }}
    .kpi-card .label {{ font-size: 12px; color: var(--gray-500); }}
    .kpi-card .value {{ font-size: 24px; font-weight: 700; color: var(--gray-900); }}
    .conclusion-card {{ padding: 12px 16px; margin-bottom: 8px; }}
    table {{ width: 100%; border-collapse: collapse; background: white;
      border: 1px solid var(--gray-200); border-radius: 8px; overflow: hidden; margin-bottom: 16px; }}
    th, td {{ padding: 10px 14px; border-top: 1px solid var(--gray-100); text-align: left; }}
    th {{ background: var(--gray-100); }}
    .pain-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }}
    .tier-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; }}
    .pain-card, .tier-card {{ padding: 14px 16px; }}
    .pain-card h4 {{ margin: 0 0 8px; }}
    .tier-card h4 {{ margin: 0 0 8px; }}
    .opportunity {{ color: var(--green); }}
    .insight {{ color: var(--gray-500); font-style: italic; margin: 8px 0 16px; }}
    .footer {{ text-align: center; font-size: 12px; color: var(--gray-500);
      margin-top: 48px; padding-top: 24px; border-top: 1px solid var(--gray-200); }}
    @media (max-width: 768px) {{
      .kpi-grid {{ grid-template-columns: repeat(2, 1fr); }}
      .pain-grid, .tier-grid {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <div class="container">
    <div class="header">
      <h1>{topic}</h1>
      <div class="subtitle">{created}</div>
      <div class="verdict-badge">{verdict}</div>
      <p>{one_line}</p>
    </div>

    <section>
      <h2>Executive Summary</h2>
      {summary_html}
    </section>

    <section>
      <h2>市场大盘</h2>
      <div class="kpi-grid">
        {kpi_html}
      </div>
    </section>

    <section>
      <h2>多维分布</h2>
      {dim_html}
    </section>

    <section>
      <h2>品牌/竞品格局</h2>
      {brand_html}
      <p class="insight">{_h(ch05.get("insight", ""))}</p>
    </section>

    <section>
      <h2>消费者痛点</h2>
      <div class="pain-grid">
        {pain_html}
      </div>
    </section>

    <section>
      <h2>机会空白</h2>
      {gap_html}
    </section>

    <section>
      <h2>策略建议</h2>
      <div class="tier-grid">
        {tier_html}
      </div>
    </section>

    <section>
      <h2>进入壁垒评估</h2>
      {barrier_html}
    </section>

    <section>
      <h2>Go/No-Go 评分卡</h2>
      {score_html}
      <p><strong>总分：{_h(total_score)} / 10 &mdash; {verdict}</strong></p>
      <p>{verdict_detail}</p>
    </section>

    <div class="footer">
      数据来源：{_h(sources)} · 生成日期：{created}
    </div>
  </div>
</body>
</html>"""


# ---------------------------------------------------------------------------
# v2 Dashboard renderer
# ---------------------------------------------------------------------------

def render_dashboard(payload: dict[str, Any], chapters: dict[str, Any], metadata: dict[str, Any]) -> str:
    """Build view-model from chapters and inject into dashboard template."""
    template_path = Path(__file__).parent.parent / "assets" / "dashboard_template.html"
    if not template_path.exists():
        raise FileNotFoundError(f"Dashboard 模板不存在: {template_path}")
    template = template_path.read_text(encoding="utf-8")

    view_model = _build_dashboard_view_model(payload, chapters, metadata)
    data_json = json.dumps(view_model, ensure_ascii=True, indent=None)

    page_title = metadata.get("topic", "Product Research Dashboard")
    from datetime import datetime
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    report_filename = f'{metadata.get("date", "")}_{metadata.get("site", "")}_{metadata.get("keyword", "")}'
    source_files = ", ".join(metadata.get("data_sources", []))

    result = template.replace("__PAGE_TITLE__", _h(page_title))
    result = result.replace("__GENERATED_AT__", generated_at)
    result = result.replace("__REPORT_FILENAME__", _h(report_filename))
    result = result.replace("__SOURCE_FILES__", _h(source_files))
    result = result.replace("__DATA_JSON__", data_json)

    return result


def _parse_share_to_number(share_str: str) -> float:
    """Parse '33.8%' → 33.8, or pass through if already numeric."""
    if isinstance(share_str, (int, float)):
        return float(share_str)
    cleaned = str(share_str).replace("%", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _build_dashboard_view_model(payload: dict[str, Any], chapters: dict[str, Any], metadata: dict[str, Any]) -> dict[str, Any]:
    """Build the DATA object consumed by the dashboard template JS.

    IMPORTANT: The dashboard template uses specific key names (often Chinese)
    and specific data shapes. This function must output EXACTLY what the
    template's init() / render*() functions expect. See dashboard_template.html
    script section for the authoritative field names.
    """
    ch01 = chapters.get("ch01_executive_summary", {})
    ch02 = chapters.get("ch02_market_overview", {})
    ch03 = chapters.get("ch03_dimension_distribution", {})
    ch04 = chapters.get("ch04_cross_analysis", {})
    ch05 = chapters.get("ch05_competitor_brands", {})
    ch06 = chapters.get("ch06_voc_pain_points", {})
    ch07 = chapters.get("ch07_opportunity_gaps", {})
    ch08 = chapters.get("ch08_strategic_recommendations", {})
    ch09 = chapters.get("ch09_barriers_go_nogo", {})

    # ── Hero section ──────────────────────────────────────────────
    conclusions = ch01.get("conclusions", [])
    hero_insights = [c.get("data_point", "") + " → " + c.get("meaning", "") for c in conclusions[:5]]
    verdict = ch09.get("verdict", ch01.get("go_nogo_verdict", ""))
    total_score = ch09.get("total_score", 0)

    focus_items = []
    avoid_items = []
    for t in ch08.get("tiers", [])[:3]:
        focus_items.append(f'{t.get("tier", "")}: {t.get("differentiation", "")}')
    for r in ch09.get("risks", [])[:3]:
        avoid_items.append(r.get("risk", ""))

    hero = {
        "title": metadata.get("topic", ""),
        "subtitle": ch01.get("one_line_summary", ""),
        "insights": hero_insights,
        "verdict": verdict,
        "score": total_score,
        "focus": focus_items,
        "avoid": avoid_items,
    }

    # ── Rings / market overview KPIs ──────────────────────────────
    rings = []
    for k in ch02.get("kpis", [])[:4]:
        rings.append({
            "name": k.get("name", ""),
            "display": k.get("value", ""),
            "value": _extract_number(k.get("value", "")),
            "hint": k.get("note", k.get("source", "")),
        })

    # Market overview metric band: template reads item["指标"], item["数值"], item["说明"]
    market_overview = []
    for k in ch02.get("kpis", []):
        market_overview.append({
            "指标": k.get("name", ""),
            "数值": k.get("value", ""),
            "说明": k.get("note", k.get("source", "")),
        })

    # ── Brands ────────────────────────────────────────────────────
    # Template: renderBarList("brandChart", DATA.brands, {labelKey:"label", valueKey:"share"})
    # secondary reads item.sales, item.count
    brands = []
    for b in ch05.get("brand_landscape", []):
        brands.append({
            "label": b.get("brand", ""),
            "share": _parse_share_to_number(b.get("market_share", "")),
            "count": b.get("product_count", 0),
            "sales": _extract_number(str(b.get("monthly_sales", b.get("product_count", 0)))),
            "avgPrice": b.get("avg_price", 0),
            "strategy": b.get("strategy", ""),
            "role": b.get("role", ""),
        })

    # ── Keywords ──────────────────────────────────────────────────
    # Template reads: point["关键词"], point["维度"], point["月搜索量"],
    #   point["CPC"], point["搜索结果竞品数"], point["旺季"]
    keywords = []
    for k in ch02.get("keyword_comparison", []):
        keywords.append({
            "关键词": k.get("keyword", ""),
            "维度": k.get("segment", ""),
            "月搜索量": k.get("monthly_search", k.get("search_volume_rank", 0)),
            "CPC": k.get("cpc", 0),
            "搜索结果竞品数": k.get("competitors", 0),
            "旺季": k.get("peak_season", ""),
        })

    # ── Distributions ─────────────────────────────────────────────
    # Template: Object.keys(DATA.distributions) → dimension names
    #   DATA.distributions[dimName] → array of items with {label, count, sales, revenue, ...}
    distributions: dict[str, list[dict[str, Any]]] = {}
    for dim in ch03.get("dimensions", []):
        dim_name = dim.get("name", "")
        items = []
        for d in dim.get("distribution", []):
            items.append({
                "label": d.get("value", ""),
                "count": d.get("product_count", 0),
                "share": d.get("product_share", ""),
                "sales": d.get("monthly_sales", 0),
                "revenue": d.get("monthly_revenue", 0),
                "avg_price": d.get("avg_price", 0),
                "salesShare": d.get("sales_share", ""),
            })
        if dim_name:
            distributions[dim_name] = items

    # ── Matrices (heatmap) ────────────────────────────────────────
    # Template renderHeatmap expects: matrix.name, matrix.rows[], matrix.cols[],
    #   matrix.rowLabel, matrix.colLabel, matrix.cells[{row, col, count, sales}]
    matrices = []
    for m in ch04.get("matrices", []):
        cells = []
        row_set: list[str] = []
        col_set: list[str] = []
        for d in m.get("data", []):
            r = d.get("dim1_value", "")
            c = d.get("dim2_value", "")
            cells.append({"row": r, "col": c, "count": d.get("count", 0), "sales": d.get("sales", 0)})
            if r and r not in row_set:
                row_set.append(r)
            if c and c not in col_set:
                col_set.append(c)
        # Also include gap cells so heatmap shows them
        for g in m.get("gaps", []):
            r = g.get("dim1_value", "")
            c = g.get("dim2_value", "")
            if r and r not in row_set:
                row_set.append(r)
            if c and c not in col_set:
                col_set.append(c)

        dim1 = m.get("dim1", "")
        dim2 = m.get("dim2", "")
        matrices.append({
            "name": f"{dim1} × {dim2}",
            "dim1": dim1,
            "dim2": dim2,
            "rowLabel": dim1,
            "colLabel": dim2,
            "rows": row_set,
            "cols": col_set,
            "cells": cells,
            "gaps": m.get("gaps", []),
            "findings": m.get("findings", ""),
        })

    # ── Gaps (opportunity) ────────────────────────────────────────
    # Template reads: item["交叉维度"], item["状态"], item["组合"], item["产品数"], item["月销量"]
    gaps = []
    for o in ch07.get("opportunities", []):
        gaps.append({
            "交叉维度": o.get("source", ""),
            "状态": f'优先级 #{o.get("rank", 0)}',
            "组合": o.get("description", ""),
            "产品数": "—",
            "月销量": o.get("weighted_score", 0) * 1000,
            "action": o.get("action", ""),
        })

    # ── Pain points ───────────────────────────────────────────────
    # Template: item["痛点维度"], item["命中次数"], item["代表吐槽"]
    pain_points = []
    for dim in ch06.get("by_dimension", []):
        top_pain = ""
        freq = 0
        pps = dim.get("pain_points", [])
        if pps:
            top_pain = pps[0].get("description", "")
            freq = pps[0].get("frequency", 0)
        pain_points.append({
            "痛点维度": dim.get("dimension", ""),
            "命中次数": freq,
            "代表吐槽": top_pain,
            "opportunity": dim.get("opportunity", ""),
            "solution": dim.get("solution", ""),
        })

    # ── Scorecard ─────────────────────────────────────────────────
    # Template: row["维度"], row["得分"], row["满分"], row["说明"]
    scorecard = []
    for s in ch09.get("go_nogo_scorecard", []):
        scorecard.append({
            "维度": s.get("dimension", ""),
            "得分": s.get("score", 0),
            "满分": 10,
            "说明": s.get("rationale", ""),
            "权重": s.get("weight", 0),
            "加权分": s.get("weighted_score", 0),
        })
    # Append total row
    scorecard.append({
        "维度": "总分",
        "得分": total_score,
        "满分": 10,
        "说明": verdict,
        "权重": 100,
        "加权分": total_score,
    })

    # ── Competitors (for scatter + table) ─────────────────────────
    # Template: point.brand, point.type, point.price, point.sales, point.reviews, point.coverage
    competitors = []
    for c in ch05.get("competitor_selection_logic", []):
        competitors.append({
            "brand": c.get("brand", ""),
            "asin": c.get("asin", ""),
            "type": c.get("type", ""),
            "price": c.get("price", 0),
            "sales": c.get("monthly_sales", 0),
            "reviews": c.get("reviews", 0),
            "coverage": c.get("covered_dimensions", ""),
        })

    # ── New product trend ─────────────────────────────────────────
    # Attempt to build from ch03 or standalone data; gracefully empty if unavailable.
    new_product_trend: list[dict[str, Any]] = []
    new_product_buckets: list[dict[str, Any]] = []
    new_product_top: list[dict[str, Any]] = []

    # ── Price histogram (from distributions) ─────────────────────
    price_histogram = distributions.get("价格带", [])

    # ── Potential products (from ch07 opportunities) ─────────────
    potential_products = []
    for opp in ch07.get("opportunities", []):
        potential_products.append({
            "brand": f'Tier {opp.get("rank", "")}',
            "potential": opp.get("weighted_score", 0),
            "asin": "",
            "sales": 0,
            "reviews": 0,
            "price": 0,
            "reason": opp.get("description", ""),
        })

    # ── Top products scatter + low review runners (from excel_sheets) ──
    excel_sheets = payload.get("excel_sheets", {})
    top100_rows = excel_sheets.get("类目销量Top100_明细", [])
    top_products_scatter: list[dict[str, Any]] = []
    low_review_runners: list[dict[str, Any]] = []
    for row in top100_rows[:30]:
        sales = _extract_number(str(row.get("月销量", 0)))
        reviews = _extract_number(str(row.get("评论数", 0)))
        price = _extract_number(str(row.get("价格", 0)))
        item = {"brand": row.get("品牌", ""), "sales": sales, "reviews": reviews, "price": price}
        top_products_scatter.append(item)
        if sales > 1000 and reviews < 500:
            low_review_runners.append({**item, "asin": row.get("ASIN", "")})

    # ── Procurement reference (from ch08) ────────────────────────
    procurement: list[dict[str, Any]] = []
    for item in ch08.get("procurement_reference", []):
        procurement.append({
            "label": item.get("label", ""),
            "min": item.get("min", 0),
            "median": item.get("median", 0),
            "max": item.get("max", 0),
        })

    return {
        "hero": hero,
        "rings": rings,
        "marketOverview": market_overview,
        "brands": brands,
        "keywords": keywords,
        "distributions": distributions,
        "matrices": matrices,
        "gaps": gaps,
        "painPoints": pain_points,
        "scorecard": scorecard,
        "competitors": competitors,
        "newProductTrend": new_product_trend,
        "newProductBuckets": new_product_buckets,
        "newProductTop": new_product_top,
        "newProductScatter": [],
        "priceHistogram": price_histogram,
        "topProductsScatter": top_products_scatter,
        "lowReviewRunners": low_review_runners,
        "dimensionDiscovery": {},
        "procurement": procurement,
        "potentialProducts": potential_products,
        "page_title": metadata.get("topic", ""),
    }


def _extract_number(text: str) -> float:
    """Extract first numeric value from a string like '$5.2M' or '55,497件'."""
    cleaned = str(text).replace(",", "").replace("$", "").replace("¥", "").replace("€", "")
    m = re.search(r"[\d.]+", cleaned)
    if not m:
        return 0
    val = float(m.group())
    if "M" in str(text) or "m" in str(text):
        val *= 1_000_000
    elif "K" in str(text) or "k" in str(text):
        val *= 1_000
    return val


# ---------------------------------------------------------------------------
# Excel writer (shared)
# ---------------------------------------------------------------------------

def write_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            continue
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        for row_idx, row in enumerate(rows, start=2):
            for col, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col, value=row.get(header, ""))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_json_artifacts(output_dir: Path, artifacts: dict[str, Any]) -> None:
    for filename, data in artifacts.items():
        target = output_dir / filename
        target.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# v1 legacy helpers
# ---------------------------------------------------------------------------

def ensure_markdown(frontmatter: str, body: str) -> str:
    if body.startswith("---\n"):
        return body
    return frontmatter + body.lstrip()


# ---------------------------------------------------------------------------
# Generate
# ---------------------------------------------------------------------------

def generate(payload: dict[str, Any], root: Path) -> dict[str, Path]:
    metadata = payload["metadata"]
    paths = build_paths(metadata, root)
    paths["dir"].mkdir(parents=True, exist_ok=True)

    if _is_v2(payload):
        chapters = payload["chapters"]
        # MD from template
        markdown = render_markdown(chapters, metadata)
        paths["md"].write_text(markdown, encoding="utf-8")
        # HTML from template
        html_content = render_html_report(chapters, metadata)
        paths["html"].write_text(html_content, encoding="utf-8")
        # Dashboard
        try:
            dashboard_content = render_dashboard(payload, chapters, metadata)
            paths["dashboard"].write_text(dashboard_content, encoding="utf-8")
        except FileNotFoundError:
            pass  # Dashboard template not available — skip
    else:
        # v1 legacy path
        markdown = ensure_markdown(render_frontmatter(metadata), payload["report_markdown"])
        paths["md"].write_text(markdown, encoding="utf-8")
        paths["html"].write_text(payload["report_html"], encoding="utf-8")

    write_xlsx(paths["xlsx"], payload["excel_sheets"])

    artifacts = payload.get("artifacts", {})
    if not isinstance(artifacts, dict):
        raise ValueError("artifacts 必须是 object")
    write_json_artifacts(paths["dir"], artifacts)
    (paths["dir"] / "excel_sheets_data.json").write_text(
        json.dumps(payload["excel_sheets"], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Save full payload for downstream tools
    (paths["dir"] / "unified_payload.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return paths


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_frontmatter(md_path: Path) -> list[str]:
    content = md_path.read_text(encoding="utf-8")
    errors = []
    if not content.startswith("---\n"):
        errors.append("Markdown 缺少 YAML 头")
        return errors
    for field in ["created:", "topic:", "type:", "data_sources:"]:
        if field not in content.split("---\n", 2)[1]:
            errors.append(f"Markdown YAML 头缺少字段: {field[:-1]}")
    return errors


def validate_html(html_path: Path) -> list[str]:
    content = html_path.read_text(encoding="utf-8")
    errors = []
    if "<html" not in content.lower():
        errors.append("HTML 文件缺少 <html>")
    if "<body" not in content.lower():
        errors.append("HTML 文件缺少 <body>")
    if "数据来源" not in content:
        errors.append('HTML Footer 缺少"数据来源"')
    if "Go / No-Go" not in content and "Go/No-Go" not in content and "Go-NoGo" not in content:
        errors.append('HTML 缺少"Go / No-Go评分卡"区块')
    if "进入壁垒" not in content:
        errors.append('HTML 缺少"进入壁垒评估"区块')
    if "品牌/竞品格局" not in content and "品牌格局" not in content and "竞品格局" not in content:
        errors.append('HTML 缺少"品牌/竞品格局"区块')
    return errors


def validate_xlsx(xlsx_path: Path, expected_sheets: dict[str, Any]) -> list[str]:
    errors = []
    wb = load_workbook(xlsx_path, read_only=True)
    sheet_names = wb.sheetnames
    if not sheet_names:
        errors.append("Excel 没有任何 Sheet")
        return errors
    if sheet_names[0] != "数据来源说明":
        errors.append('Excel 首 Sheet 不是"数据来源说明"')
    for sheet in REQUIRED_SHEETS:
        if sheet not in sheet_names:
            errors.append(f"Excel 缺少必选 Sheet: {sheet}")
    for sheet in expected_sheets.keys():
        if sheet not in sheet_names:
            errors.append(f"Excel 缺少 payload 中声明的 Sheet: {sheet}")
    return errors


def validate_v2_chapters(chapters: dict[str, Any]) -> list[str]:
    """Validate v2 chapters structure and required fields."""
    errors = []
    for key in CHAPTER_KEYS:
        if key not in chapters:
            errors.append(f"chapters 缺少: {key}")

    ch01 = chapters.get("ch01_executive_summary", {})
    conclusions = ch01.get("conclusions", [])
    if len(conclusions) < 3:
        errors.append(f"ch01 conclusions 数量不足: {len(conclusions)}（需 ≥ 3）")
    if not ch01.get("go_nogo_verdict"):
        errors.append("ch01 缺少 go_nogo_verdict")

    ch02 = chapters.get("ch02_market_overview", {})
    kpis = ch02.get("kpis", [])
    if len(kpis) < 5:
        errors.append(f"ch02 kpis 数量不足: {len(kpis)}（需 ≥ 5）")
    if not ch02.get("insight"):
        errors.append("ch02 缺少 insight")

    ch03 = chapters.get("ch03_dimension_distribution", {})
    dims = ch03.get("dimensions", [])
    if len(dims) < 2:
        errors.append(f"ch03 dimensions 数量不足: {len(dims)}（需 ≥ 2）")
    for d in dims:
        if not d.get("insight"):
            errors.append(f"ch03 维度「{d.get('name', '?')}」缺少 insight")

    ch04 = chapters.get("ch04_cross_analysis", {})
    for m in ch04.get("matrices", []):
        if not m.get("findings"):
            errors.append(f"ch04 交叉分析「{m.get('dim1', '?')}×{m.get('dim2', '?')}」缺少 findings")

    ch05 = chapters.get("ch05_competitor_brands", {})
    if not ch05.get("insight"):
        errors.append("ch05 缺少 insight")

    ch07 = chapters.get("ch07_opportunity_gaps", {})
    if not ch07.get("opportunities"):
        errors.append("ch07 缺少 opportunities")

    ch09 = chapters.get("ch09_barriers_go_nogo", {})
    verdict = ch09.get("verdict", "")
    if verdict not in ("GO", "CONDITIONAL GO", "HOLD", "NO-GO", ""):
        errors.append(f"ch09 verdict 非法值: {verdict}")
    if not ch09.get("go_nogo_scorecard"):
        errors.append("ch09 缺少 go_nogo_scorecard")

    return errors


def validate_insight_quality(md_path: Path) -> list[str]:
    """检查报告的洞察质量，确保不是纯数据堆砌。"""
    content = md_path.read_text(encoding="utf-8")
    lines = content.split("\n")
    warnings = []

    if "Executive Summary" not in content and "executive summary" not in content.lower():
        warnings.append("洞察质量：缺少 Executive Summary")

    insight_count = content.count("关键洞察") + content.count("关键发现")
    if insight_count < 3:
        warnings.append(f"洞察质量：「关键洞察」/「关键发现」仅出现 {insight_count} 次（需 ≥ 3）")

    opportunity_count = content.count("品牌机会") + content.count("产品方案") + content.count("差异化")
    if opportunity_count < 3:
        warnings.append(f"洞察质量：品牌机会/产品方案/差异化相关内容仅出现 {opportunity_count} 次（需 ≥ 3）")

    has_tier = bool(re.search(r"Tier\s*[123]", content, re.IGNORECASE))
    if not has_tier:
        warnings.append("洞察质量：缺少产品矩阵（Tier 1/2/3）")
    else:
        tier_section = content[content.lower().find("tier"):] if "tier" in content.lower() else ""
        if tier_section and re.search(r"待确认|待定|待 Zach 确认|TBD", tier_section[:500]):
            warnings.append("洞察质量：产品矩阵含「待确认」/「待定」占位，必须给出具体规格")

    if len(lines) < 250:
        warnings.append(f"洞察质量：报告仅 {len(lines)} 行（需 ≥ 250），分析深度可能不足")

    if "进入壁垒评估" not in content and "进入壁垒" not in content:
        warnings.append('洞察质量：缺少「进入壁垒评估」章节')

    if "Go / No-Go" not in content and "Go/No-Go" not in content and "Go-NoGo" not in content:
        warnings.append('洞察质量：缺少「Go / No-Go评分卡」章节')

    if "品牌/竞品格局" not in content and "品牌格局" not in content and "竞品格局" not in content:
        warnings.append('洞察质量：缺少「品牌/竞品格局」章节')

    return warnings


def validate_payload_structure(payload: dict[str, Any]) -> list[str]:
    errors = []
    sheets = payload["excel_sheets"]
    if list(sheets.keys())[0] != "数据来源说明":
        errors.append('excel_sheets 第一个 key 必须为"数据来源说明"')
    for sheet in REQUIRED_SHEETS:
        if sheet not in sheets:
            errors.append(f"payload.excel_sheets 缺少必选 Sheet: {sheet}")
    return errors


def validate_paths(metadata: dict[str, Any], paths: dict[str, Path]) -> list[str]:
    errors = []
    expected_dir = Path("工作成果") / "brands" / metadata["brand"] / "市场调研" / metadata["category"] / metadata["version"]
    if not str(paths["dir"]).endswith(str(expected_dir)):
        errors.append(f"输出目录不符合规范: {paths['dir']}")
    base_name = build_base_name(metadata)
    for key in ["md", "html", "xlsx"]:
        if not paths[key].name.startswith(base_name):
            errors.append(f"{key.upper()} 文件名不符合 date_site_keyword 规范")
    return errors


def validate(payload: dict[str, Any], root: Path) -> tuple[dict[str, Path], list[str]]:
    metadata = payload["metadata"]
    paths = build_paths(metadata, root)
    errors = []

    # v2 chapters validation
    if _is_v2(payload):
        errors.extend(validate_v2_chapters(payload["chapters"]))

    errors.extend(validate_payload_structure(payload))
    errors.extend(validate_paths(metadata, paths))

    for key in ["md", "html", "xlsx"]:
        if not paths[key].exists():
            errors.append(f"缺少交付文件: {paths[key].name}")

    if not errors:
        errors.extend(validate_frontmatter(paths["md"]))
        errors.extend(validate_html(paths["html"]))
        errors.extend(validate_xlsx(paths["xlsx"], payload["excel_sheets"]))
        errors.extend(validate_insight_quality(paths["md"]))

    # Dashboard existence check for v2
    if _is_v2(payload) and not errors:
        if not paths["dashboard"].exists():
            errors.append(f"缺少 Dashboard 文件: {paths['dashboard'].name}")

    return paths, errors


# ---------------------------------------------------------------------------
# CLI entry
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="生成并校验选品调研交付物（v1 三件套 / v2 四件套）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
例子：
  # 从脚本所在目录运行
  python render_deliverables.py all --input ../../payload.json

  # 从任意目录运行（使用绝对路径）
  python /path/to/render_deliverables.py all --input C:\\absolute\\path\\payload.json

  # 指定输出根目录
  python render_deliverables.py all --input payload.json --root C:\\output\\root
        """
    )
    parser.add_argument("command", choices=["generate", "validate", "all"])
    parser.add_argument("--input", "-i", required=True, help="统一 JSON 数据包路径（支持相对或绝对路径）")
    parser.add_argument("--root", default=".", help="仓库根目录，默认当前目录")
    args = parser.parse_args()

    try:
        # 确保路径是绝对路径，支持相对和绝对路径
        input_path = Path(args.input).resolve()
        if not input_path.exists():
            print(f"❌ 错误：输入文件不存在", file=sys.stderr)
            print(f"   搜索位置：{input_path}", file=sys.stderr)
            return 1

        root = Path(args.root).resolve()

        payload = read_payload(input_path)

        if args.command in {"generate", "all"}:
            paths = generate(payload, root)
            print(f"✅ 生成完成：{paths['dir']}")

        if args.command in {"validate", "all"}:
            paths, errors = validate(payload, root)
            if errors:
                print(f"❌ 校验失败：{paths['dir']}", file=sys.stderr)
                for error in errors:
                    print(f"   - {error}", file=sys.stderr)
                return 1
            print(f"✅ 校验通过：{paths['dir']}")

        return 0

    except FileNotFoundError as e:
        print(f"❌ 错误：文件不存在 - {e}", file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"❌ 错误：数据验证失败 - {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"❌ 未预期的错误 - {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
